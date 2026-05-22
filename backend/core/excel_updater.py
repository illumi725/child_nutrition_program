import logging
import openpyxl
import os

logger = logging.getLogger(__name__)

def update_excel_birthday(file_path: str, row_number: int, new_birthday_str: str) -> bool:
    """
    Updates the birthday cell for a specific row in the Form 5A Excel file.
    Row numbers here should correspond to the 1-based index in Excel.
    """
    if not os.path.exists(file_path):
        logger.error("Excel file not found: %s", file_path)
        return False
        
    try:
        wb = openpyxl.load_workbook(file_path)
        
        # Find the correct sheet
        sheet = None
        for s in wb.sheetnames:
            if '5A' in s.upper() or 'FORM 5A' in s.upper():
                sheet = wb[s]
                break
                
        if not sheet:
            sheet = wb.active
            
        # Find the birthday column index
        # We'll scan the first 30 rows to find the header row containing "LAST NAME" or "FIRST NAME"
        # and then find the column with "BIRTHDAY" or "BIRTH"
        birthday_col_idx = -1
        header_row_idx = -1
        
        for r_idx in range(1, min(31, sheet.max_row + 1)):
            row_vals = [str(sheet.cell(row=r_idx, column=c).value).upper() for c in range(1, sheet.max_column + 1)]
            row_str = " ".join(row_vals)
            
            if 'LAST NAME' in row_str or 'FIRST NAME' in row_str:
                header_row_idx = r_idx
                for c_idx, val in enumerate(row_vals):
                    if 'BIRTHDAY' in val or 'BIRTH' in val:
                        birthday_col_idx = c_idx + 1 # openpyxl is 1-indexed
                        break
                break
                
        if birthday_col_idx == -1:
            logger.error("Could not locate Birthday column in %s", file_path)
            return False
            
        # Update the cell
        cell = sheet.cell(row=row_number, column=birthday_col_idx)
        cell.value = new_birthday_str
        
        wb.save(file_path)
        return True
        
    except Exception as e:
        logger.error("Failed to update birthday in %s: %s", file_path, e, exc_info=True)
        return False

def update_excel_name(file_path: str, row_number: int, new_name_str: str) -> bool:
    """
    Updates the name cell for a specific row in the Form 5A Excel file.
    """
    if not os.path.exists(file_path):
        return False
        
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = next((wb[s] for s in wb.sheetnames if '5A' in s.upper() or 'FORM 5A' in s.upper()), wb.active)
            
        name_col_idx = -1
        
        for r_idx in range(1, min(31, sheet.max_row + 1)):
            row_vals = [str(sheet.cell(row=r_idx, column=c).value).upper() for c in range(1, sheet.max_column + 1)]
            row_str = " ".join(row_vals)
            
            if 'LAST NAME' in row_str or 'FIRST NAME' in row_str:
                for c_idx, val in enumerate(row_vals):
                    if 'NAME' in val:
                        name_col_idx = c_idx + 1
                        break
                break
                
        if name_col_idx == -1:
            return False
            
        cell = sheet.cell(row=row_number, column=name_col_idx)
        cell.value = new_name_str
        wb.save(file_path)
        return True
        
    except Exception as e:
        logger.error("Failed to update name in %s: %s", file_path, e, exc_info=True)
        return False


def delete_excel_row(file_path: str, row_number: int) -> tuple:
    """
    Permanently deletes a row from the Form 5A Excel file.
    Rows below the deleted row shift up automatically (openpyxl delete_rows).
    Returns (success: bool, error_message: str | None)
    """
    if not os.path.exists(file_path):
        return False, f"File not found: {file_path}"
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = next(
            (wb[s] for s in wb.sheetnames if '5A' in s.upper() or 'FORM 5A' in s.upper()),
            wb.active
        )
        sheet.delete_rows(row_number)
        wb.save(file_path)
        return True, None
    except Exception as e:
        return False, str(e)

